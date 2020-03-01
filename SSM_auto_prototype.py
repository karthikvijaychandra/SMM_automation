from selenium import webdriver
import csv
import urllib.request
from openpyxl import Workbook


# Definition
URLDataToWrite = []
ImageSizeDataToWrite = []
FileSize = []
i = 0

# Logic
driver = webdriver.Chrome('./chromedriver')
driver.maximize_window()
with open('URL.txt', mode='r') as URLData:
    for item in URLData:
        driver.get(item)
        URLDataToWrite.append(" ")
        URLDataToWrite.append(item.split('build/')[1].capitalize().rstrip())
        imgVar = driver.find_elements_by_tag_name('img')
        for item1 in imgVar:
            if item1.get_attribute('src'):
                URLDataToWrite.append(item1.get_attribute('src'))
            elif item1.get_attribute('data-src'):
                dataSRC = item1.get_attribute('data-src')
                URLDataToWrite.append('http://seemymachines.qburst.build/' + dataSRC)
        CssImgElements = driver.find_elements_by_xpath("//*[starts-with(@class,'icon ')]")
        for item2 in CssImgElements:
            URLDataToWrite.append(item2.value_of_css_property('background-image').split('"')[1])

driver.quit()

# Write to CSV file
'''
with open('Data_SMM.csv', 'w') as CSVFile:
    writeCSV = csv.writer(CSVFile, delimiter='\n')
    writeCSV.writerow(URLDataToWrite)'''

# Download images
'''for item in URLDataToWrite:
    if item.__contains__('base64'):
        urllib.request.urlretrieve(item, './SSM_images/'+'data'+str(i)+'.png')
        i += 1
    if item.__contains__('images'):
        urllib.request.urlretrieve(item, './SSM_images/'+item.split('images/')[1])'''

# Get file size
for item in URLDataToWrite:
    if item.__contains__('image') or item.__contains__('qburst.build'):
        URLInfo = urllib.request.urlopen(item)
        FileSize.append(int(URLInfo.info()['Content-Length'])/1000)
    else:
        FileSize.append(" ")

for item in FileSize:
    if type(item) == int or type(item) == float:
        ImageSizeDataToWrite.append(str(item) + ' KB')
    else:
        ImageSizeDataToWrite.append(" ")


# Write to excel Document
wb = Workbook()
sheet = wb.active
sheet.title = 'SeeMyMachines'
sheet['A1'] = 'Image URL'
sheet['B1'] = 'Size'

# Write URLs to Excel
for num, val in enumerate(URLDataToWrite, start=2):
    cell = sheet.cell(row=num, column=1)
    cell.value = val


# Write images size to Excel
for num, val in enumerate(ImageSizeDataToWrite, start=2):
    cell2 = sheet.cell(row=num, column=2)
    cell2.value = val

wb.save('./textexcel.xlsx')